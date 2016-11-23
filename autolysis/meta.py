'''
Generates standardised metadata for tabular datasets. Typical usage::

    from autolysis import Metadata
    info = Metadata('tests/data/x.csv')
    info = Metadata('file:///tests/data/x.csv')
'''
from __future__ import unicode_literals

import io
import os
import sys
import json
import time
import shutil
import logging
import openpyxl
import requests
import subprocess
import pandas as pd
import sqlalchemy as sa
from copy import copy
from hashlib import md5
from six import text_type, string_types
from functools import wraps
from itertools import islice, chain
from orderedattrdict import AttrDict

from six.moves.urllib_parse import urlparse
from .config import DATA_DIR

# tqdm imports colorama. colorama disables IPython TAB completion in Py2.
# https://github.com/tartley/colorama/issues/92
# Avoid importing tqdm in Py2 IPython
try:
    __IPYTHON__
except NameError:
    __IPYTHON__ = False
if sys.version_info[0] == 2 and __IPYTHON__:
    def tqdm(x, *args, **kwargs):
        return x
else:
    from tqdm import tqdm

OK = 200                    # HTTP status code
seconds_per_day = 86400     # Number of seconds in a day
expiry_days = 1             # Number of days after which URLs expiry
_excel_template = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'metadata-template.xlsx')


def metadata(source, tables=None, root=None, merge=True, **kwargs):
    '''
    Return the metadata for the selected source as a Meta.
    '''
    if root is None:
        root = os.path.join(DATA_DIR, '.metadata')
    if not os.path.exists(root):
        os.makedirs(root)

    # Extract base metadata along with commands to extract the data
    tree = Meta(source=source)
    scheme = urlparse(source).scheme
    if os.path.exists(source) or scheme in {'file'}:
        tree.update(metadata_file(source, root, tables))
    elif scheme in {'http', 'https', 'ftp'}:
        target = filename(source, root)
        fetch(source, target)
        tree.update(metadata_file(target, root, tables))
    else:
        tree.format = 'sql'
        tree.update(metadata_sql(source, tables))

    # Extract sub-datasets
    dataset_list = list(datasets(tree))
    for node in tqdm(dataset_list, disable=kwargs.get('tqdm_disable')):
        cmd = node.get('command', [None])
        if cmd[0] in _preview_command:
            try:
                data = _preview_command[cmd[0]](*cmd[1:])
                node.update(metadata_frame(data, **kwargs))
            except Exception as e:
                node['error'] = text_type(e)
                logging.exception('Unable to load %s', ':'.join(cmd[1:]))

    # Merge column metadata of common datasets
    if merge:
        for node in datasets(tree):
            if 'datasets' in node:
                sign_lookup = {}
                for data in node.datasets.values():
                    if 'columns' in data:
                        sign = tuple(col.name for col in data.columns.values())
                        if sign in sign_lookup:
                            data.columns = Column(see=sign_lookup[sign].name)
                        else:
                            sign_lookup[sign] = data

    return tree


def metadata_sql(source, tables=None):
    '''
    Returns metadata for a SQLAlchemy source URL for a subset of tables
    '''
    try:
        engine = sa.create_engine(source)       # noqa: encoding kills Py2.7
    except sa.exc.ArgumentError:
        raise NotImplementedError('Cannot process source %s' % source)
    # Get a list of database source URLs
    inspector = sa.inspect(engine)
    if not engine.url.database:
        source = source.rstrip('/') + '/'
        tree = Meta(datasets=Datasets())
        for schema in inspector.get_schema_names():
            subtree = tree.datasets[schema] = Meta(name=schema, format='sql', datasets=Datasets())
            # Special case: For PostgreSQL, treat 'public' as ''
            if schema == 'public' and engine.url.drivername.startswith('postgresql'):
                schema = None
            for table in (tables or inspector.get_table_names(schema)):
                subtree.datasets[table] = Meta([
                    ('name', table), ('format', 'table'),
                    ('command', ['sql', table, source + (schema or '')]),
                ])
    else:
        tree = Meta(datasets=Datasets())
        for table in (tables or inspector.get_table_names()):
            tree.datasets[table] = Meta([
                ('name', table), ('format', 'table'),
                ('command', ['sql', table, source]),
            ])
    return tree


def metadata_file(path, root, tables=None):
    '''
    Returns the metadata for a file. There are 3 types of file formats:

    1. Archives (7z, zip, rar, tar) / compressed (xz, bzip2, gzip). Decompress and process
    2. Database (sqlite3, hdf5, xls, xlsx). Process each table/sheet as a sub-dataset
    3. Data (csv, json). Process directly
    '''
    tree = Meta()
    format = guess_format(path)
    if format is not None:
        tree.format = format

    if format == 'dir':
        tree.datasets = Datasets()
        for base, dirs, files in os.walk(path):
            for filename in files:
                source = os.path.join(base, filename)
                name = os.path.relpath(source, path)
                tree.datasets[name] = submeta = Meta(name=name, source=source)
                try:
                    submeta.update(metadata_file(source, root, tables))
                except Exception as e:
                    submeta['error'] = text_type(e)
                    logging.exception('Unable to get metadata for %s', source)
    elif format in {'7z', 'zip', 'rar', 'tar', 'xz', 'gz', 'bz2'}:
        tree.datasets = Datasets()
        for name, source in unzip_files(path, root, format):
            tree.datasets[name] = submeta = Meta(name=name)
            try:
                submeta.update(metadata_file(source, root, tables))
            except Exception as e:
                submeta['error'] = text_type(e)
                logging.exception('Unable to get metadata for %s', source)
    elif format == 'sqlite3':
        tree.update(metadata_sql('sqlite:///' + path, tables))
    elif format in {'hdf5', 'xls', 'xlsx'}:
        if format == 'hdf5':
            store = pd.HDFStore(path)
            # Pandas categorical data creates a /meta/values_block_<n>/meta table
            # Ignore these?
            table_list = [key for key in store.keys() if 'values_block_' not in key]
            store.close()
        else:
            xls = pd.ExcelFile(path)
            table_list = xls.sheet_names
            format = 'xlsx'
        tree.datasets = Datasets()
        for table in table_list:
            if tables is None or table in tables:
                tree.datasets[table] = Meta([
                    ('name', table),
                    ('format', 'table'),
                    ('command', [format, path, table])
                ])
    elif format in {'csv', 'json', 'dta'}:
        tree['command'] = [format, path]
    return tree


def metadata_frame(data, top=3, preview=10, **kwargs):
    '''
    Compute the metadata for a Pandas DataFrame
    '''
    columns = Columns()
    for col, series in data.iteritems():
        meta = Column(name=text_type(col))
        meta.type_pandas = series.dtype.name
        meta.missing = int(pd.isnull(series).sum())
        meta.nunique = series.nunique()
        # TODO: Preserve order for these transformations
        meta.top = series.value_counts().head(top)
        if series.dtype.kind in {'i', 'f'}:
            meta.moments = series.dropna().describe()
        columns[meta.name] = meta
    preview = min(preview, len(data))
    result = Meta([('rows', len(data)), ('columns', columns)])
    result['head'] = data.head(preview)
    result['sample'] = data.sample(preview).sort_index()
    return result


def datasets(tree):
    yield tree
    for node in tree.get('datasets', Datasets()).values():
        for subnode in datasets(node):
            yield subnode


def unzip_files(path, root, format):
    '''
    Extract all files in the archive at path into root using format specified.
    Yield (name, path) tuples for each file in the archive.

    For multi-file archives like 7z, ZIP, RAR, TAR, the name is the relative
    path to the archived file name. For single file archives like gz, bzip2,
    xz, the name is the extracted filename.

    In all cases, the path is the full path to the extracted file.
    '''
    if format in {'7z', 'zip', 'rar', 'tar'}:
        target = filename(path, root)
        # If the target is outdated, delete it.
        # Note: for ZIP inside ZIP, the first extraction creates a new file.
        # So the inner ZIP will ALWAYS be extracted. TODO: resolve this.
        if os.path.exists(target) and os.stat(target).st_mtime < os.stat(path).st_mtime:
            shutil.rmtree(target)
        # If the target does not exist, create it
        if not os.path.exists(target):
            os.makedirs(target)
            extract_archive(path, target=target, format=format)
        # Return each file
        for base, dirs, files in os.walk(target):
            for path in files:
                # path is relative to archive base
                name = os.path.join(os.path.relpath(base, target), path)
                yield os.path.normpath(name), os.path.join(base, path)
    elif format in {'xz', 'gz', 'bz2'}:
        # Strip out the extension from the filename. That's the new name of the dataset
        name = os.path.split(path)[-1]
        if name.endswith(format):
            name = name[:-len(format) - 1]
        target = os.path.join(root, name)
        # TODO: if target is fresh, do not re-extract
        if os.path.exists(target):
            os.unlink(target)
        extract_archive(path, target=root, format=format)
        yield name, target


def extract_archive(archive, target, format):
    '''
    Extract the archive (with specified format) into target directory
    '''
    cmd = [
        '7z', 'x', '-y',                # Extract from 7-zip without prompting
        '-t' + _format_map.get(format, format),    # with specified format
        '-o%s' % target,                # into target directory
        archive,                        # from the source archive
    ]
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    returncode = proc.wait()
    if returncode != 0:
        raise subprocess.CalledProcessError(returncode, cmd, proc.stdout.read())

    # 7zip adds a PaxHeaders.nnn subdirectory for tar files. Ignore that.
    if format == 'tar':
        for root, dirs, files in os.walk(target):
            for dirname in dirs:
                if dirname.lower().startswith('paxheaders.'):
                    shutil.rmtree(os.path.join(root, dirname), ignore_errors=False)


def filename(url, root, path=None):
    '''Return unique filename from url (urlhash-name.ext) under root'''
    if path is None:
        path = urlparse(url).path
    name, ext = os.path.splitext(os.path.split(path)[-1])
    filename = '{urlhash}-{name}{ext}'.format(
        name=name, ext=ext, urlhash=md5(url.encode('utf-8')).hexdigest()[:10])
    return os.path.abspath(os.path.join(root, filename))


def fetch(url, path):
    '''
    Retrieves the HTTP or FTP url and saves it into path, unless path is newer than expiry_days.
    Returns the path
    '''
    now = time.time()
    if os.path.exists(path) and os.stat(path).st_mtime > now - expiry_days * seconds_per_day:
        return path
    r = requests.get(url)
    if r.status_code == OK and len(r.content):
        with io.open(path, 'wb') as handle:
            handle.write(r.content)
    # TODO: in case of ANY failure, raise an exception


def guess_format(path, ignore_ext=False):
    '''
    Returns file format for data files based on the file extension or signature.
    File signatures / magic are from http://www.garykessler.net/library/file_sigs.html
    '''
    base, ext = os.path.splitext(path)
    ext = ext.lower()[1:]
    if ext and ext in _ext_map and not ignore_ext:
        return _ext_map[ext]

    # Guess based on signature
    if os.path.isdir(path):     # Directories are allowed, and have a 'dir' format
        return 'dir'
    signature_length = 20       # It's enough to read ~20 bytes to identify the file
    with io.open(path, 'rb') as handle:
        head = handle.read(signature_length)
    if head.startswith(b'7z\xbc\xaf\x27\x1c'):
        return '7z'
    if head.startswith(b'PK') and head[2:4] in {b'\x03\x04', b'\x05\x06', b'\x07\x08'}:
        from zipfile import ZipFile
        for filename in ZipFile(path).namelist():
            if filename.startswith('xl/'):
                return 'xlsx'
        return 'zip'
    if head.startswith(b'Rar!\x1A\x07\x00'):
        return 'rar'
    if head.startswith(b'ustar') or head.startswith(b'./PaxHeaders'):
        return 'tar'
    if head.startswith(b'\xfd7zXZ\x00'):
        return 'xz'
    if head.startswith(b'\x1f\x8b'):
        return 'gz'
    if head.startswith(b'BZh'):
        return 'bz2'
    if head.startswith(b'SQLite format 3\x00'):
        return 'sqlite3'
    if head.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
        return 'xls'
    if head.startswith(b'\x89\x48\x44\x46\x0d\x0a\x1a\x0a'):
        return 'hdf5'
    # http://www.stata.com/help.cgi?dta_115 and ?dta_118
    if head.startswith(b'<stata_dta>'):
        return 'dta'
    if (b'\x71' <= head[0:1] <= b'\x73' and head[1:2] in {b'\x01', b'\x02'}
            and head[2:4] == b'\x01\x00'):
        return 'dta'
    try:
        read_json(path)
        return 'json'
    except (ValueError, pd.core.common.PandasError):
        pass
    try:
        # Since almost any file matches CSV, ensure it has at least 1 row, 2 cols
        result = next(read_csv_encoded(path, chunksize=1000))
        if len(result) >= 1 and len(result.columns) >= 2:
            return 'csv'
    except Exception:
        pass


def read_encoded(method):
    '''
    Read a CSV file via Pandas ``read_csv`` with different encodings. By default,
    try CP-1252 and then UTF-8. Set ``encodings=('encoding', ...)`` to change the
    list of encodings to try.
    '''
    def wrapped(*args, **kwargs):
        for encoding in kwargs.pop('encodings', ('cp1252', 'utf-8', 'utf-16-le', 'utf-16-be')):
            try:
                kwargs['encoding'] = encoding
                result = method(*args, **kwargs)
                if 'chunksize' in kwargs:
                    peek = next(result)
                    return chain([peek], result)
                else:
                    return result
            except Exception as e:
                pass
        raise e

    return wrapped


def read_json(*args, **kwargs):
    '''
    Read a CSV file via Pandas ``read_json``. If the JSON has only 1 row, return
    a DataFrame with 1 row.
    '''
    try:
        return pd.read_json(*args, **kwargs)
    except ValueError as e:
        if e.args[0] == 'If using all scalar values, you must pass an index':
            kwargs['typ'] = 'series'
            return pd.read_json(*args, **kwargs).to_frame()
        raise


class MetaDict(AttrDict):
    '''
    Base class for all metadata structures
    '''
    def _repr_pretty_(self, p, cycle):
        '''Printing the object in IPython prints str(object)'''
        lines = self.__str__().split('\n')
        for line in lines[:-1]:
            p.text(line)
            p.break_()
        p.text(lines[-1])

    def to_json(self, drop={'command', 'head', 'sample'}, **kwargs):
        '''Return an ordered JSON string containing the metadata'''
        # Create a copy and remove keys that need not be exported
        meta = copy(self)
        for key in list(meta):
            if key in drop:
                meta.pop(key)
        # Set default JSON export properties and export as JSON
        import json
        kwargs.setdefault('cls', PandasEncoder)
        kwargs.setdefault('indent', 4)
        return json.dumps(meta, **kwargs)

    def to_dict(self):
        '''Return a plain dict version of the metadata'''
        return json.loads(self.to_json(indent=0))

    def to_yaml(self, **kwargs):
        '''Return a YAML string containing the metadata'''
        import yaml
        import orderedattrdict.yamlutils            # noqa. Imported to preserve order in YAML
        kwargs.setdefault('default_flow_style', False)
        data = json.loads(self.to_json(indent=0), object_pairs_hook=AttrDict)
        return yaml.safe_dump(data, **kwargs)

    def to_text(self, rows=100):
        '''Return a hierarchical text list of datasets and column names'''
        return self.__str__(rows=100, deep=True)

    def to_markdown(self):
        '''Return a Markdown list of lists of datasets and column names'''
        return self.__str__(rows=100, deep=True, template='markdown')


_templates = {
    'text': {
        'datasets': '{name:s} ({fmt:s}) {size:d} datasets',
        'dataset': '{name:s} ({fmt:s}) {rows:,d}{more:s} rows {cols:,d} cols',
        'no-dataset': '{name:s} ({fmt:s}). No datasets/rows detected',
        'more': '...',
        'section': '{name:s}:',
        'attr': '{key:s}: {val:s}',
        'column': '{name:s} ({type:s}): {top:s}',
    },
    'markdown': {
        'datasets': '- **{name:s}** ({fmt:s}) {size:d} datasets',
        'dataset': '- **{name:s}** ({fmt:s}) {rows:,d}{more:s} rows {cols:,d} cols',
        'no-dataset': '- **{name:s}** ({fmt:s}). No datasets/rows detected',
        'more': '- ...',
        'section': '- **{name:s}**:',
        'attr': '- *{key:}*: {val:s}',
        'column': '- **{name:s}** ({type:s}): {top:s}',
    },
}


class Meta(MetaDict):
    def __str__(self, rows=100, deep=False, template=None):
        tmpl = _templates.get(template, _templates['text'])
        name = self.get('name', self.get('source', ''))
        format = self.get('format', 'format?')
        result = []
        if 'datasets' in self:
            result.append(tmpl['datasets'].format(name=name, fmt=format, size=len(self.datasets)))
            if rows:
                result += ['    ' + data
                           for data in self.datasets.__str__(rows, deep, template).split('\n')]
        elif 'rows' in self and 'columns' in self:
            result.append(tmpl['dataset'].format(
                name=name, fmt=format,
                rows=self.rows,
                more='+' if self.rows == MAX_ROWS else '',
                cols=len(self.columns)))
            if rows:
                result += ['    ' +
                           col for col in self.columns.__str__(rows, deep, template).split('\n')]
        else:
            result.append(tmpl['no-dataset'].format(name=name, fmt=format))
        return '\n'.join(result)

    def data(self):
        cmd = self.get('command', [None])
        if cmd[0] in _read_command:
            return _read_command[cmd[0]](*cmd[1:])

    @staticmethod
    def _expression(expression, meta):
        # Find the value of the Python expression in the cell.
        # If the expression contains a [*] we treat it as a list enumeration.
        expression = expression.strip().lstrip('#').strip()
        try:
            if '[*]' not in expression:
                # ['x'][0]['z'] => meta['x'][0]['z']
                return eval('meta' + expression)
            else:
                # ['columns'][*]['name'] => [meta['columns'][i]['name']]
                prefix, postfix = expression.split('[*]', 2)
                return [eval('item' + postfix) for item in eval('meta' + prefix + '.values()')]
        except Exception as e:
            logging.exception('Error in expression %s', expression)
            return None

    def to_excel(self, target_file, template=None):
        # Loop through the excel sheet and find every cell that begins with //
        # mapping[sheetname, row, column] has the contents of the cell
        meta = self.to_dict()
        book = openpyxl.load_workbook(template or _excel_template)
        for sheetname in book.sheetnames:
            sheet = book.get_sheet_by_name(sheetname)
            for i, row in enumerate(sheet.rows):
                for j, cell in enumerate(row):
                    if isinstance(cell.value, string_types) and cell.value.startswith('#'):
                        value = self._expression(cell.value, meta)
                        # If the expression returns a list, write values one below another
                        if isinstance(value, list):
                            for index in range(len(value)):
                                sheet.cell(row=i + 1 + index, column=j + 1).value = value[index]
                        # If expression returns a single value, use it as is
                        elif value is not None:
                            cell.value = value
        book.save(target_file)

    def to_powerpoint(self):
        pass


class CollectionMixin(object):
    def __str__(self, rows=100, deep=False, template=None):
        tmpl = _templates.get(template, _templates['text'])
        result = []
        indent = '    ' if deep and self._show_name else ''
        for name, col in islice(self.items(), rows):
            if deep and self._show_name:
                result.append(tmpl['section'].format(name=name))
            lines = col.__str__(rows if deep else 0, deep, template)
            result += [indent + line for line in lines.split('\n')]
        if len(self) > rows:
            result.append(tmpl['more'])
        return '\n'.join(result)


class Datasets(MetaDict, CollectionMixin):
    _show_name = False


class Columns(MetaDict, CollectionMixin):
    _show_name = True


class Column(MetaDict):
    def __str__(self, rows=100, deep=False, template=None):
        '''
        Display a column as follows::

            name: column name       # column name as-is
            type_pandas: int64      # or whatever Pandas type it is
            missing: 30             # Number of missing values
            nunique: 100            # Number of unique values
            top:
                 first: 100
                second:  95
                 third:   1
        '''
        tmpl = _templates.get(template, _templates['text'])
        if rows > 0:
            result = []
            for key in ['name', 'type_pandas', 'missing', 'nunique']:
                if key in self:
                    result.append(tmpl['attr'].format(key=key, val=text_type(self[key])))
            for section, dtype in [('top', 'int'), ('moments', 'float')]:
                if section in self:
                    result.append(tmpl['section'].format(name=section))
                    if template == 'text':
                        max_val = max(self[section].values)
                        fmt_key = '{:>%ds}' % max(len('%s' % key) for key in self[section].index)
                    else:
                        max_val = 0
                        fmt_key = '{:s}'
                    if dtype == 'int':
                        fmt_val = '{:%dd}' % len('{:d}'.format(max_val))
                    elif dtype == 'float':
                        fmt_val = '{:%d.2f}' % len('{:.2f}'.format(max_val))
                    else:
                        fmt_val = '{:r}'
                    fmt_key, fmt_val = fmt_key.format, fmt_val.format
                    for key, val in self[section].iteritems():
                        result.append('    ' + tmpl['attr'].format(
                            key=fmt_key(text_type(key)), val=fmt_val(val)))
            return '\n'.join(result[:rows])
        else:
            top = ', '.join('%s' % val for val in self.top.keys())
            return tmpl['column'].format(name=self.name, type=self.type_pandas, top=top)


class PandasEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, pd.Series):
            result = json.loads(obj.to_json(orient='split'))
            return AttrDict(zip(result['index'], result['data']))
        elif isinstance(obj, pd.DataFrame):
            return json.loads(obj.to_json(orient='split'))
        return json.JSONEncoder.default(self, obj)


def chunked(method, chunksize):
    '''Wrap Pandas read_* methods with a specified chunksize and returns the first chunk'''
    @wraps(method)
    def wrapped(*args, **kwargs):
        kwargs['chunksize'] = chunksize
        try:
            for result in method(*args, **kwargs):
                return result
        # pd.read_hdf5 raises a TypeError on fixed format files since it cannot
        # read them as an iterator. Load them fully instead
        except TypeError:
            del kwargs['chunksize']
            return method(*args, **kwargs)
    return wrapped


MAX_ROWS = 10000
read_csv_encoded = read_encoded(pd.read_csv)
read_stata_encoded = read_encoded(pd.read_stata)
_preview_command = {
    'csv': chunked(read_csv_encoded, MAX_ROWS),
    'dta': chunked(read_stata_encoded, MAX_ROWS),
    'json': read_json,
    'sql': chunked(pd.read_sql_table, MAX_ROWS),
    'xlsx': pd.read_excel,
    'hdf5': chunked(pd.read_hdf, MAX_ROWS),
}

_read_command = {
    'csv': read_csv_encoded,
    'dta': read_stata_encoded,
    'json': read_json,
    'sql': pd.read_sql_table,
    'xlsx': pd.read_excel,
    'hdf5': pd.read_hdf,
}

# List of known extensions and the types they map to
_ext_map = {
    '7z': '7z',
    '7zip': '7z',
    'bz2': 'bz2',
    'csv': 'csv',
    'db': 'sqlite3',
    'dta': 'dta',
    'gz': 'gz',
    'h5': 'hdf5',
    'hdf5': 'hdf5',
    'json': 'json',
    'rar': 'rar',
    'sqlite3': 'sqlite3',
    'tar': 'tar',
    'xls': 'xls',
    'xlsx': 'xlsx',
    'xz': 'xz',
    'zip': 'zip',
}

_format_map = {
    'gz': 'gzip',
    'bz2': 'bzip2',
}
